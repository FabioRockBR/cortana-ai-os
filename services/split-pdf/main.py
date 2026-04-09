"""
Google Cloud Function: Process Large Files for RAG Ingestion

Full pipeline: Download file → Process → Embed → Store in Qdrant.
Triggered by n8n when a file exceeds the 20MB size limit.

Supported file types:
  - PDF/text: Download → Extract Text → Chunk → batchEmbedContents → Qdrant
  - Image:    Download → base64 → embedContent (inline_data) → Qdrant (1 point)
  - Video:    Download → base64 → embedContent (inline_data, ≤50MB) or chunked → Qdrant

Environment variables:
    GOOGLE_API_KEY: Google AI Studio API key for Gemini Embedding 2
    QDRANT_URL: Qdrant Cloud URL (e.g. https://...qdrant.io)
    QDRANT_API_KEY: Qdrant API key

Deploy:
    cd tools/cloud_function_split_pdf
    gcloud functions deploy split-pdf \
        --gen2 \
        --runtime python311 \
        --trigger-http \
        --allow-unauthenticated \
        --memory 2Gi \
        --timeout 540s \
        --entry-point process_file \
        --region europe-west1 \
        --service-account split-pdf-sa@n8n-2026-486511.iam.gserviceaccount.com \
        --set-env-vars "GOOGLE_API_KEY=...,QDRANT_URL=...,QDRANT_API_KEY=..."
"""

import base64
import hashlib
import io
import logging
import os
import re
import tempfile
import time
import uuid
from datetime import datetime, timezone

import flask
import functions_framework
import requests
import fitz  # pymupdf
from google.auth import default
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload


# --- Config ---
SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
DEFAULT_COLLECTION = 'aircraft_maintenance_kb_v2'   # legacy / general fallback
EMBEDDING_MODEL = 'gemini-embedding-2-preview'
CHUNK_SIZE = 2000
CHUNK_OVERLAP = 400
EMBEDDING_BATCH = 50    # Gemini batchEmbedContents limit
QDRANT_BATCH = 100
VIDEO_CHUNK_BYTES = 50 * 1024 * 1024   # 50MB ≈ 128s compressed video (Gemini limit)

GEMINI_BASE = 'https://generativelanguage.googleapis.com/v1beta'

# Fleet → Qdrant collection mapping
# New filename convention: {FLEET}_{DOCTYPE}_ATA{NN}_{Description}.pdf
# e.g. B737NG_AMM_ATA27_FlightControls_Rev45.pdf
FLEET_PATTERNS = [
    # 737 Classic must come before generic 737 NG patterns
    (r'B737[\s_-]?CLASSIC|737[\s_-]?CLASSIC|B737CL\b|B737[\s_-]?(?:100|200|300|400|500)\b', 'B737-Classic', 'b737_classic_kb'),
    (r'B737[\s_-]?NG|737[\s_-]?NG|737NG|B737[\s_-]?(?:600|700|800|900)\b',                  'B737-NG',      'b737_ng_kb'),
    (r'\bB757\b|\b757\b',                                                              'B757',         'b757_kb'),
    (r'\bB767\b|\b767\b',                                                              'B767',         'b767_kb'),
    (r'\bA32[0-9]\b|\bA31[89]\b',                                                     'A320',         'a320_kb'),
    (r'\bB747\b|\b747\b',                                                              'B747',         'b747_kb'),
    (r'\bB777\b|\b777\b',                                                              'B777',         'b777_kb'),
]


# --- Google Drive ---

def get_drive_service():
    creds, _ = default(scopes=SCOPES)
    return build('drive', 'v3', credentials=creds)


def download_file_from_url(url: str, suffix: str = '.bin') -> str:
    """Download from a public HTTPS URL to a temp file. Raises on non-HTTPS."""
    if not url.startswith('https://'):
        raise ValueError('Only HTTPS URLs are accepted')
    resp = requests.get(url, timeout=60, stream=True)
    resp.raise_for_status()
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    for chunk in resp.iter_content(chunk_size=65536):
        tmp.write(chunk)
    tmp.close()
    return tmp.name


def download_file(service, file_id, suffix='.pdf'):
    req = service.files().get_media(fileId=file_id)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    fh = io.FileIO(tmp.name, 'wb')
    downloader = MediaIoBaseDownload(fh, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    fh.close()
    return tmp.name


# --- Text Processing ---

def _detect_header_row(rows):
    """Return (header_list, data_start_index) for a table's rows.

    A header row is the first row where the majority of non-empty cells are
    predominantly alphabetic (not just short numbers).  We scan the first
    three rows so multi-row column headings are handled.
    """
    for idx, row in enumerate(rows[:3]):
        cells = [str(c).strip() if c else '' for c in row]
        non_empty = [c for c in cells if c]
        if not non_empty:
            continue
        # Count cells that look like labels (contain at least one letter)
        label_count = sum(1 for c in non_empty if re.search(r'[A-Za-z]', c))
        if label_count >= max(1, len(non_empty) // 2):
            return cells, idx + 1
    return None, 0


def _extract_tables(page, vertical_strategy='lines', horizontal_strategy='lines'):
    """Run find_tables with the given strategies; return [] on failure."""
    try:
        result = page.find_tables(
            vertical_strategy=vertical_strategy,
            horizontal_strategy=horizontal_strategy,
        )
        return result.tables if result else []
    except Exception:
        return []


def _format_table_page(page):
    """Extract a page's tables as structured key:value text.

    Tries line-based detection first (works for bordered tables), then falls
    back to text-alignment-based detection (works for borderless MEL-style
    columns aligned by whitespace).

    Every data row is emitted as "Header1: val | Header2: val | ..." so that
    all columns for a single MEL/AMM item stay in the same text chunk.
    Non-table regions on the same page are appended as plain text below.

    Returns an empty string if no tables are found (caller falls back to
    plain page.get_text()).
    """
    # Strategy 1: line-bordered tables (AMM, IPC, revision pages)
    tables = _extract_tables(page, 'lines', 'lines')

    # Strategy 2: borderless column-aligned tables (MEL operational pages)
    if not tables:
        tables = _extract_tables(page, 'text', 'lines')

    if not tables:
        return ''

    table_bboxes = []
    structured_parts = []

    for table in tables:
        table_bboxes.append(table.bbox)
        rows = table.extract()
        if not rows:
            continue

        header, data_start = _detect_header_row(rows)

        for row in rows[data_start:]:
            cells = [str(c).strip() if c else '' for c in row]
            if not any(cells):
                continue
            if header:
                parts = [f"{h}: {v}" for h, v in zip(header, cells) if h and v]
            else:
                parts = [v for v in cells if v]
            if parts:
                structured_parts.append(' | '.join(parts))

    if not structured_parts:
        return ''

    # Grab plain text from regions outside detected tables
    clip_text = []
    for block in page.get_text('blocks'):
        bx0, by0, bx1, by1, text, *_ = block
        block_rect = fitz.Rect(bx0, by0, bx1, by1)
        in_table = any(block_rect.intersects(fitz.Rect(tb)) for tb in table_bboxes)
        if not in_table and text.strip():
            clip_text.append(text.strip())

    result = '\n'.join(structured_parts)
    if clip_text:
        result += '\n\n' + '\n'.join(clip_text)
    return result


def extract_text(pdf_path):
    doc = fitz.open(pdf_path)
    total_pages = len(doc)
    all_text = []
    for page in doc:
        # Try table-aware extraction first; fall back to plain text
        text = _format_table_page(page)
        if not text:
            text = page.get_text()
        if text.strip():
            all_text.append(text)
    doc.close()
    return '\n\n'.join(all_text), total_pages


def parse_excel_structured(file_path):
    """Return Excel workbook as structured JSON (sheets → headers + rows dict list).

    Used by the Maintenance Report Importer to extract raw row data for Airtable
    mapping. Unlike extract_excel_text() which formats rows as prose for RAG,
    this returns the raw cell values keyed by column header.

    Returns:
        {
            "sheets": [{"name": str, "headers": [str], "rows": [{col: val, ...}]}],
            "total_rows": int
        }
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows_raw = list(ws.iter_rows(values_only=True))
        if not rows_raw:
            continue
        # First non-empty row is treated as headers
        headers = [
            str(c).strip() if c is not None else f'Col{i}'
            for i, c in enumerate(rows_raw[0])
        ]
        rows = []
        for row in rows_raw[1:]:
            cells = [str(c).strip() if c is not None else '' for c in row]
            if not any(cells):
                continue
            rows.append(dict(zip(headers, cells)))
        if rows:
            sheets.append({'name': name, 'headers': headers, 'rows': rows})
    wb.close()
    return {'sheets': sheets, 'total_rows': sum(len(s['rows']) for s in sheets)}


def extract_excel_text(file_path):
    """Extract text from an Excel workbook (.xlsx / .xls).

    Each sheet is rendered as a section header followed by rows formatted as
    "Col1: val | Col2: val | ..." (same key:value style as PDF table extraction).
    Reuses _detect_header_row so merged/labelled header rows are handled cleanly.

    Returns (text, num_sheets).
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    num_sheets = len(wb.sheetnames)
    all_parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header, data_start = _detect_header_row(rows)
        sheet_parts = [f'=== Sheet: {sheet_name} ===']
        for row in rows[data_start:]:
            cells = [str(c).strip() if c is not None else '' for c in row]
            if not any(cells):
                continue
            if header:
                parts = [f"{h}: {v}" for h, v in zip(header, cells) if h and v]
            else:
                parts = [v for v in cells if v]
            if parts:
                sheet_parts.append(' | '.join(parts))
        if len(sheet_parts) > 1:
            all_parts.append('\n'.join(sheet_parts))
    wb.close()
    return '\n\n'.join(all_parts), num_sheets


def split_text(text):
    """Split text into overlapping chunks at clean break points."""
    if not text.strip():
        return []
    if len(text) <= CHUNK_SIZE:
        return [text.strip()]

    chunks = []
    start = 0
    while start < len(text):
        end = min(start + CHUNK_SIZE, len(text))

        if end < len(text):
            # Find a clean break: paragraph > newline > space
            bp = text.rfind('\n\n', start + CHUNK_SIZE // 2, end)
            if bp == -1:
                bp = text.rfind('\n', start + CHUNK_SIZE // 2, end)
            if bp == -1:
                bp = text.rfind(' ', start + CHUNK_SIZE // 2, end)
            if bp != -1:
                end = bp + 1

        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)

        # Advance with overlap, or finish if we've reached the end
        start = end - CHUNK_OVERLAP if end < len(text) else len(text)

    return chunks


def detect_metadata(file_name, folder_path=''):
    """Extract fleet, ATA chapter, and document type from filename + folder path.

    New naming convention (preferred):
        B737NG_AMM_ATA27_FlightControls_Rev45.pdf
        B737Classic_WDM_ATA23_Navigation.pdf
        B757_IPC_ATA32_LandingGear.pdf

    Swiftair internal convention (___doc_code suffix):
        {ATA}___{docCode}.pdf   e.g. 21___052.pdf = ATA 21, B737 Classic
        FM___{docCode}.pdf      Fleet Manual for that fleet
        {alpha}___{docCode}.pdf Other Swiftair docs

        Doc code → fleet mapping:
          052 = B737 Classic, 089 = B757, 103 = B737 NG
          139 = B757 (alternate series), 026/016/102 = Swiftair institutional

    Legacy support: 21___092.pdf → ATA 21 detected from leading digits
    Fleet can also be inferred from the Google Drive folder path.

    Returns: (ata_chapter, doc_type, fleet, collection)
    """
    # Swiftair doc_code → (fleet_label, collection)
    SWT_DOC_CODES = {
        '052': ('B737-Classic', 'b737_classic_kb'),
        '089': ('B757',         'b757_kb'),
        '103': ('B737-NG',      'b737_ng_kb'),
        '139': ('B757',         'b757_kb'),
    }

    # Combine filename + folder path for detection; upper-case both
    search_text = (folder_path + ' ' + file_name).upper()
    name_no_ext = os.path.splitext(file_name)[0].upper()

    # --- Fleet detection (folder path takes priority over filename) ---
    fleet = 'GENERAL'
    collection = DEFAULT_COLLECTION
    for pattern, fleet_name, coll in FLEET_PATTERNS:
        if re.search(pattern, search_text):
            fleet = fleet_name
            collection = coll
            break

    # --- Swiftair doc_code fallback (only if no fleet detected yet) ---
    # Matches patterns like "21___052", "FM___089", "AA___016"
    if fleet == 'GENERAL':
        doc_code_match = re.search(r'___(\d{3})', name_no_ext)
        if doc_code_match:
            code = doc_code_match.group(1)
            if code in SWT_DOC_CODES:
                fleet, collection = SWT_DOC_CODES[code]

    # --- ATA chapter ---
    # Try "ATA27", "ATA-27", "ATA 27" first
    ata_match = re.search(r'ATA[\s_-]?(\d{2})', search_text)
    if not ata_match:
        # Legacy: leading two-digit number followed by separator ("21___092")
        ata_match = re.search(r'^(\d{2})[\s_-]', name_no_ext)
    ata_chapter = ata_match.group(1) if ata_match else 'unknown'

    # --- Document type (order matters — more specific first) ---
    doc_type = 'GENERAL'
    for pattern, val in [
        # Swiftair-specific types (check before generic AMM/CMM)
        (r'\bMOE\b',                                           'MOE'),
        (r'F-PO-MNT|F_PO_MNT',                               'PROCEDURE'),
        (r'\bQUIP\b',                                          'QUALITY_PROCEDURE'),
        (r'M-SMS|M_SMS',                                       'SMS_PROCEDURE'),
        (r'M-ORG|M_ORG',                                       'MANAGEMENT_MANUAL'),
        (r'DAILY[\s_-]CHECK|WEEKLY[\s_-]CHECK|INSPECTION[\s_-]CHECK|48[\s_-]HR[\s_-]CHECK', 'INSPECTION_CHECK'),
        (r'GOOD[\s_-]PRACTICES[\s_-]GUIDE|GOOD[\s_-]PRACTICES', 'GUIDE'),
        (r'\bNI___\b|\bNOTICE\b',                             'NOTICE'),
        (r'\bIN___\b|\bINSTRUCTION\b',                        'INSTRUCTION'),
        (r'\bSMS\b',                                           'SMS_MANUAL'),
        # Standard types
        (r'\bWDM\b|\bWIRING\b',                               'WDM'),
        (r'\bIPC\b',                                           'IPC'),
        (r'\bINCIDENT\b|\bPHOTO\b|\bVIDEO\b',                'INCIDENT'),
        (r'\bAMM\b',                                           'AMM'),
        (r'\bCMM\b',                                           'CMM'),
        (r'\bSRM\b',                                           'SRM'),
        (r'\bMEL\b',                                           'MEL'),
        (r'\bTSM\b',                                           'TSM'),
        (r'SERVICE[\s_-]BULLETIN|\bSB\b',                     'SB'),
        (r'AIRWORTHINESS|\bAD\b',                             'AD'),
        (r'FAULT[\s_-]ISOLATION|\bFIM\b',                     'FI'),
    ]:
        if re.search(pattern, search_text):
            doc_type = val
            break

    return ata_chapter, doc_type, fleet, collection


# --- Gemini Embeddings ---

def generate_embeddings(chunks, api_key):
    """Generate text embeddings in batches via Gemini batchEmbedContents API.

    Uses taskType=RETRIEVAL_DOCUMENT so the model optimizes vectors for
    storage-side retrieval (paired with RETRIEVAL_QUERY at query time in Cortana).
    """
    all_embeddings = []
    clean_chunks = [c.replace('\n', ' ') for c in chunks]
    model_path = f'models/{EMBEDDING_MODEL}'

    for i in range(0, len(clean_chunks), EMBEDDING_BATCH):
        batch = clean_chunks[i:i + EMBEDDING_BATCH]
        requests_payload = [
            {
                'model': model_path,
                'content': {'parts': [{'text': text}]},
                'taskType': 'RETRIEVAL_DOCUMENT',
            }
            for text in batch
        ]
        for _attempt in range(6):
            resp = requests.post(
                f'{GEMINI_BASE}/{model_path}:batchEmbedContents',
                headers={'Content-Type': 'application/json'},
                params={'key': api_key},
                json={'requests': requests_payload},
                timeout=120,
            )
            if resp.status_code == 429:
                wait = 10 * (2 ** _attempt)
                logging.warning('Gemini 429, retrying in %ds (attempt %d/6)', wait, _attempt + 1)
                time.sleep(wait)
                continue
            break
        if not resp.ok:
            raise RuntimeError(
                f'Gemini batchEmbedContents {resp.status_code}: {resp.text[:500]}'
            )
        embeddings = [e['values'] for e in resp.json()['embeddings']]
        all_embeddings.extend(embeddings)

    return all_embeddings


def embed_media_inline(data_bytes, mime_type, api_key):
    """Embed a single image or video chunk via Gemini embedContent (inline_data).

    Images and videos cannot use batchEmbedContents — they require individual
    embedContent calls with the file encoded as base64 inline_data.

    Returns a list of 3072 floats.
    """
    model_path = f'models/{EMBEDDING_MODEL}'
    resp = requests.post(
        f'{GEMINI_BASE}/{model_path}:embedContent',
        headers={'Content-Type': 'application/json'},
        params={'key': api_key},
        json={
            'model': model_path,
            'content': {
                'parts': [{
                    'inline_data': {
                        'mime_type': mime_type,
                        'data': base64.b64encode(data_bytes).decode('utf-8'),
                    }
                }]
            },
            'taskType': 'RETRIEVAL_DOCUMENT',
        },
        timeout=120,
    )
    if not resp.ok:
        raise RuntimeError(
            f'Gemini embedContent {resp.status_code}: {resp.text[:500]}'
        )
    return resp.json()['embedding']['values']


def embed_image(file_path, mime_type, api_key):
    """Embed a single image file. Returns (embeddings_list, num_points)."""
    with open(file_path, 'rb') as f:
        data = f.read()
    vector = embed_media_inline(data, mime_type, api_key)
    return [vector], 1


def embed_pdf_pages_as_images(pdf_path, api_key):
    """Render each PDF page as a PNG image and embed it via Gemini inline_data.

    Used for WDM and IPC documents where visual content (wiring diagrams,
    connector pin-outs, illustrated parts breakdowns) would be lost during
    text extraction.  Each page becomes one Qdrant point.

    Renders at 150 DPI (matrix 150/72 ≈ 2.08x) — sharp enough for technical
    diagrams without oversizing the Gemini payload.

    Returns (embeddings_list, num_pages).
    """
    doc = fitz.open(pdf_path)
    embeddings = []
    mat = fitz.Matrix(150 / 72, 150 / 72)
    for page in doc:
        pix = page.get_pixmap(matrix=mat, alpha=False)
        png_bytes = pix.tobytes('png')
        vector = embed_media_inline(png_bytes, 'image/png', api_key)
        embeddings.append(vector)
        pix = None  # release pixmap memory immediately
    doc.close()
    return embeddings, len(embeddings)


def embed_video(file_path, mime_type, api_key):
    """Embed a video file, chunking if >50MB (Gemini's ~128s inline_data limit).

    Each chunk produces one Qdrant point representing a temporal segment.
    Returns (embeddings_list, num_chunks).
    """
    with open(file_path, 'rb') as f:
        data = f.read()

    if len(data) <= VIDEO_CHUNK_BYTES:
        vector = embed_media_inline(data, mime_type, api_key)
        return [vector], 1

    # Split into byte chunks — each embedded as a separate temporal segment
    embeddings = []
    total = len(data)
    offset = 0
    while offset < total:
        chunk_bytes = data[offset:offset + VIDEO_CHUNK_BYTES]
        vector = embed_media_inline(chunk_bytes, mime_type, api_key)
        embeddings.append(vector)
        offset += VIDEO_CHUNK_BYTES

    return embeddings, len(embeddings)


# --- Qdrant ---

def delete_file_vectors(file_id, qdrant_url, qdrant_key, collection=DEFAULT_COLLECTION):
    """Delete existing vectors for a file (dedup on re-ingestion)."""
    try:
        requests.post(
            f'{qdrant_url}/collections/{collection}/points/delete',
            headers={'api-key': qdrant_key, 'Content-Type': 'application/json'},
            json={
                'filter': {
                    'must': [{
                        'key': 'metadata.source_file_id',
                        'match': {'value': file_id},
                    }]
                }
            },
            timeout=30,
        )
    except Exception:
        pass  # Non-critical — continue even if cleanup fails


def upsert_vectors(points, qdrant_url, qdrant_key, collection=DEFAULT_COLLECTION):
    """Upsert points to Qdrant in batches."""
    total = 0
    for i in range(0, len(points), QDRANT_BATCH):
        batch = points[i:i + QDRANT_BATCH]
        resp = requests.put(
            f'{qdrant_url}/collections/{collection}/points',
            headers={'api-key': qdrant_key, 'Content-Type': 'application/json'},
            json={'points': batch},
            timeout=60,
        )
        resp.raise_for_status()
        total += len(batch)
    return total


# --- Entry Point ---

@functions_framework.http
def process_file(request):
    """
    Full RAG ingestion pipeline for large files (PDF, image, video).

    Expects JSON body (one of fileId / fileUrl / pdfBase64 required):
    {
        "fileId":    "google-drive-file-id",   // Google Drive file
        "fileUrl":   "https://...",            // public HTTPS URL (e.g. Catbox image)
        "pdfBase64": "...",                    // base64-encoded PDF bytes
        "fileName":  "document.pdf",
        "mimeType":  "application/pdf"         // optional, defaults to application/pdf
    }

    Returns:
    {
        "status": "success",
        "fileName": "...",
        "modality": "text|image|video",
        "totalPages": 757,         // PDF only
        "totalChunks": 2384,
        "vectorsStored": 2384
    }
    """
    # Auth check — require X-CF-Token header when CF_AUTH_TOKEN env var is set
    expected_token = os.environ.get('CF_AUTH_TOKEN', '')
    if expected_token:
        provided_token = request.headers.get('X-CF-Token', '')
        if provided_token != expected_token:
            return flask.jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json(silent=True)
    if not data:
        return flask.jsonify({'error': 'No JSON body provided'}), 400

    # List mode: return all files in a Drive folder (for batch ingestion scripts)
    list_folder_id = data.get('listFolderId')
    if list_folder_id:
        service = get_drive_service()
        q = f"'{list_folder_id}' in parents and trashed=false"
        result = service.files().list(q=q, fields='files(id,name,mimeType)', pageSize=200).execute()
        files = result.get('files', [])
        return flask.jsonify({'files': files, 'total': len(files)})

    # Parse-only mode: returns structured JSON rows, does NOT ingest to Qdrant.
    # Used by the Maintenance Report Importer workflow to extract task rows from
    # an Excel file (e.g. open maintenance tasks report) for Airtable mapping.
    if data.get('action') == 'parseExcel':
        file_url = data.get('fileUrl')
        if not file_url:
            return flask.jsonify({'error': 'fileUrl required for parseExcel action'}), 400
        try:
            tmp_path = download_file_from_url(file_url, suffix='.xlsx')
            result = parse_excel_structured(tmp_path)
            os.unlink(tmp_path)
            return flask.jsonify(result), 200
        except Exception as exc:
            logging.exception('parseExcel failed')
            return flask.jsonify({'error': str(exc)}), 500

    file_id    = data.get('fileId')
    file_url   = data.get('fileUrl')
    pdf_base64 = data.get('pdfBase64')
    file_name  = data.get('fileName', 'document.pdf')
    mime_type  = data.get('mimeType', 'application/pdf')
    # Optional: caller can pass the Drive folder path to help fleet detection
    # e.g. folderPath = "RAG - Maintenance Manuals/B737-NG/AMM"
    folder_path = data.get('folderPath', '')
    # Optional: caller can override the target collection directly
    collection_override = data.get('collection', '')

    if not file_id and not file_url and not pdf_base64:
        return flask.jsonify({'error': 'Provide fileId, fileUrl, or pdfBase64'}), 400

    if file_id:
        if not re.match(r'^[a-zA-Z0-9_-]{10,}$', str(file_id)):
            return flask.jsonify({'error': 'Invalid fileId format'}), 400
        source_id = file_id
    elif file_url:
        source_id = 'url_' + hashlib.md5(file_url.encode()).hexdigest()
    else:
        source_id = 'b64_' + hashlib.md5(pdf_base64[:500].encode()).hexdigest()

    # Reject unknown MIME types rather than silently treating them as PDF
    _allowed_mime_prefixes = ('application/pdf', 'text/', 'image/', 'video/')
    _allowed_mime_exact = {
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
        'application/vnd.ms-excel',                                            # .xls
    }
    if not any(mime_type.startswith(p) for p in _allowed_mime_prefixes) and mime_type not in _allowed_mime_exact:
        return flask.jsonify({'error': f'Unsupported MIME type: {mime_type}'}), 400

    google_api_key = os.environ.get('GOOGLE_API_KEY')
    qdrant_url = os.environ.get('QDRANT_URL', '').rstrip('/')
    qdrant_key = os.environ.get('QDRANT_API_KEY')

    if not all([google_api_key, qdrant_url, qdrant_key]):
        return flask.jsonify({'error': 'Missing environment variables: GOOGLE_API_KEY, QDRANT_URL, QDRANT_API_KEY'}), 500

    # Determine file suffix for temp file
    ext_map = {
        'image/jpeg': '.jpg', 'image/png': '.png', 'image/webp': '.webp',
        'image/heic': '.heic', 'image/gif': '.gif',
        'video/mp4': '.mp4', 'video/quicktime': '.mov',
        'video/x-msvideo': '.avi', 'video/webm': '.webm',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': '.xlsx',
        'application/vnd.ms-excel': '.xls',
    }
    suffix = ext_map.get(mime_type, '.pdf')

    ata_chapter, doc_type, fleet, collection = detect_metadata(file_name, folder_path)
    if collection_override:
        collection = collection_override
    now = datetime.now(timezone.utc).isoformat()
    modality = 'image' if mime_type.startswith('image/') else ('video' if mime_type.startswith('video/') else 'text')

    try:
        if file_id:
            service = get_drive_service()
            file_path = download_file(service, file_id, suffix=suffix)
        elif file_url:
            file_path = download_file_from_url(file_url, suffix=suffix)
        else:
            pdf_bytes = base64.b64decode(pdf_base64)
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
            tmp.write(pdf_bytes)
            tmp.close()
            file_path = tmp.name

        # --- Image path ---
        if mime_type.startswith('image/'):
            embeddings, num_points = embed_image(file_path, mime_type, google_api_key)
            os.unlink(file_path)

            delete_file_vectors(source_id, qdrant_url, qdrant_key, collection)

            points = [{
                'id': str(uuid.uuid4()),
                'vector': embeddings[0],
                'payload': {
                    'content': f'[IMAGE] {file_name}',
                    'metadata': {
                        'source_file': file_name,
                        'source_file_id': source_id,
                        'modality': 'image',
                        'mime_type': mime_type,
                        'ata_chapter': ata_chapter,
                        'document_type': doc_type,
                        'fleet': fleet,
                        'collection': collection,
                        'chunk_index': 0,
                        'total_chunks': 1,
                        'ingested_at': now,
                    },
                },
            }]
            vectors_stored = upsert_vectors(points, qdrant_url, qdrant_key, collection)

            return flask.jsonify({
                'status': 'success',
                'fileName': file_name,
                'modality': 'image',
                'totalChunks': 1,
                'vectorsStored': vectors_stored,
            }), 200

        # --- Video path ---
        if mime_type.startswith('video/'):
            embeddings, num_chunks = embed_video(file_path, mime_type, google_api_key)
            os.unlink(file_path)

            delete_file_vectors(source_id, qdrant_url, qdrant_key, collection)

            points = []
            for i, vector in enumerate(embeddings):
                points.append({
                    'id': str(uuid.uuid4()),
                    'vector': vector,
                    'payload': {
                        'content': f'[VIDEO] {file_name} (segment {i + 1}/{num_chunks})',
                        'metadata': {
                            'source_file': file_name,
                            'source_file_id': source_id,
                            'modality': 'video',
                            'mime_type': mime_type,
                            'ata_chapter': ata_chapter,
                            'document_type': doc_type,
                            'fleet': fleet,
                            'collection': collection,
                            'chunk_index': i,
                            'total_chunks': num_chunks,
                            'ingested_at': now,
                        },
                    },
                })
            vectors_stored = upsert_vectors(points, qdrant_url, qdrant_key, collection)

            return flask.jsonify({
                'status': 'success',
                'fileName': file_name,
                'modality': 'video',
                'totalChunks': num_chunks,
                'vectorsStored': vectors_stored,
            }), 200

        # --- Excel path (.xlsx / .xls) ---
        _excel_mimes = {
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel',
        }
        if mime_type in _excel_mimes:
            text, num_sheets = extract_excel_text(file_path)
            os.unlink(file_path)

            if not text.strip():
                return flask.jsonify({'error': 'No text extracted from Excel file', 'fileName': file_name}), 400

            chunks = split_text(text)
            del text
            embeddings = generate_embeddings(chunks, google_api_key)
            delete_file_vectors(source_id, qdrant_url, qdrant_key, collection)

            points = []
            for i, (chunk, vector) in enumerate(zip(chunks, embeddings)):
                points.append({
                    'id': str(uuid.uuid4()),
                    'vector': vector,
                    'payload': {
                        'content': chunk,
                        'metadata': {
                            'source_file': file_name,
                            'source_file_id': source_id,
                            'modality': 'text',
                            'mime_type': mime_type,
                            'ata_chapter': ata_chapter,
                            'document_type': doc_type,
                            'fleet': fleet,
                            'collection': collection,
                            'chunk_index': i,
                            'total_chunks': len(chunks),
                            'ingested_at': now,
                        },
                    },
                })
            vectors_stored = upsert_vectors(points, qdrant_url, qdrant_key, collection)

            return flask.jsonify({
                'status': 'success',
                'fileName': file_name,
                'modality': 'text',
                'fleet': fleet,
                'collection': collection,
                'totalSheets': num_sheets,
                'totalChunks': len(chunks),
                'vectorsStored': vectors_stored,
            }), 200

        # --- WDM / IPC: page-as-image path (preserves wiring diagrams, pin-outs, illustrated parts) ---
        if doc_type in ('WDM', 'IPC'):
            embeddings, num_pages = embed_pdf_pages_as_images(file_path, google_api_key)
            os.unlink(file_path)

            delete_file_vectors(source_id, qdrant_url, qdrant_key, collection)

            points = []
            for i, vector in enumerate(embeddings):
                points.append({
                    'id': str(uuid.uuid4()),
                    'vector': vector,
                    'payload': {
                        'content': f'[{doc_type}_PAGE] {file_name} page {i + 1}/{num_pages}',
                        'metadata': {
                            'source_file': file_name,
                            'source_file_id': source_id,
                            'modality': 'image',
                            'mime_type': 'image/png',
                            'ata_chapter': ata_chapter,
                            'document_type': doc_type,
                            'fleet': fleet,
                            'collection': collection,
                            'chunk_index': i,
                            'total_chunks': num_pages,
                            'page_number': i + 1,
                            'ingested_at': now,
                        },
                    },
                })
            vectors_stored = upsert_vectors(points, qdrant_url, qdrant_key, collection)

            return flask.jsonify({
                'status': 'success',
                'fileName': file_name,
                'modality': 'image',
                'documentType': doc_type,
                'totalPages': num_pages,
                'vectorsStored': vectors_stored,
            }), 200

        # --- PDF / text path (default) ---
        text, total_pages = extract_text(file_path)
        os.unlink(file_path)

        if not text.strip():
            return flask.jsonify({
                'status': 'error',
                'error': 'No text extracted from PDF',
                'fileName': file_name,
            }), 200

        chunks = split_text(text)
        del text  # Free memory

        embeddings = generate_embeddings(chunks, google_api_key)

        delete_file_vectors(source_id, qdrant_url, qdrant_key, collection)

        points = []
        for i, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
            points.append({
                'id': str(uuid.uuid4()),
                'vector': embedding,
                'payload': {
                    'content': chunk,
                    'metadata': {
                        'source_file': file_name,
                        'source_file_id': source_id,
                        'modality': 'text',
                        'ata_chapter': ata_chapter,
                        'document_type': doc_type,
                        'fleet': fleet,
                        'collection': collection,
                        'chunk_index': i,
                        'total_chunks': len(chunks),
                        'ingested_at': now,
                    },
                },
            })
        del chunks, embeddings  # Free memory

        vectors_stored = upsert_vectors(points, qdrant_url, qdrant_key, collection)

        return flask.jsonify({
            'status': 'success',
            'fileName': file_name,
            'modality': 'text',
            'fleet': fleet,
            'collection': collection,
            'totalPages': total_pages,
            'totalChunks': len(points),
            'vectorsStored': vectors_stored,
        }), 200

    except Exception:
        logging.exception('Unhandled error processing %s (%s)', file_name, modality)
        return flask.jsonify({
            'status': 'error',
            'error': 'Internal server error',
            'fileName': file_name,
        }), 500


# Backward-compatible alias — old deploy commands used process_pdf as entry point
process_pdf = process_file


# ---------------------------------------------------------------------------
# NotebookLM Research Pipeline
# ---------------------------------------------------------------------------

def _summarize_mindmap(data):
    """Extract top-level node titles from a NotebookLM mind map JSON object."""
    import json as _json
    try:
        if isinstance(data, str):
            data = _json.loads(data)
        # NotebookLM returns {"mind_map": {"name": "...", "children": [...]}}
        if 'mind_map' in data:
            data = data['mind_map']
        children = data.get('children', data.get('nodes', data.get('topics', [])))
        nodes = []
        for c in children[:8]:
            if isinstance(c, dict):
                title = c.get('name', c.get('title', c.get('label', c.get('text', ''))))
                if title:
                    nodes.append(str(title))
        return '\n'.join(f'• {n}' for n in nodes) if nodes else ''
    except Exception:
        return ''


# Input limits and allowed values
_CF_MAX_TOPIC_LENGTH = 500
_CF_MAX_SOURCE_LENGTH = 2048
_CF_MAX_SOURCES = 20
_CF_ALLOWED_ARTIFACT_TYPES = frozenset({'audio', 'report', 'mindmap'})


def _cf_sanitize_topic(topic) -> str:
    """Strip control characters and enforce maximum length."""
    if not isinstance(topic, str):
        return 'Research'
    cleaned = ''.join(
        ch for ch in topic
        if ch in ('\t', '\n', ' ') or (ord(ch) >= 0x20 and ord(ch) != 0x7F)
    )
    return cleaned[:_CF_MAX_TOPIC_LENGTH].strip() or 'Research'


@functions_framework.http
def research_notebooklm(request):
    """
    Run a full NotebookLM research pipeline: create notebook, add sources,
    generate report + mind map + audio overview, upload artifacts to GCS,
    and return signed URLs.

    POST body:
        topic          (str)       — Research topic (used as notebook title, max 500 chars)
        sources        (list[str]) — HTTPS URLs or plain-text strings (max 20, https only for URLs)
        artifact_types (list[str]) — Subset of ["audio", "report", "mindmap"]
                                     Default: all three

    Returns:
        JSON {
            status:          "success" | "error",
            sources_added:   int,
            summary:         str,       # AI-generated notebook overview
            report_text:     str,       # Markdown study guide
            mindmap_json:    dict,      # Raw mind map structure
            mindmap_summary: str,       # Bullet list of top-level nodes
            mindmap_url:     str,       # GCS signed URL (1 hour)
            audio_url:       str,       # GCS signed URL (1 hour)
            audio_error:     str,       # Only present if audio failed
            error:           str        # Only present on fatal failure
        }

    Deploy:
        gcloud functions deploy research-notebooklm \\
            --gen2 --runtime python311 --trigger-http \\
            --allow-unauthenticated --memory 1Gi --timeout 540s \\
            --entry-point research_notebooklm --region europe-west1 \\
            --service-account split-pdf-sa@n8n-2026-486511.iam.gserviceaccount.com \\
            --set-secrets "NOTEBOOKLM_AUTH_JSON=notebooklm-auth-json:latest" \\
            --set-env-vars "GCS_RESEARCH_BUCKET=research-artifacts-n8n-2026"
    """
    import asyncio
    import json as _json
    import traceback as _tb
    from datetime import timedelta

    data = request.get_json(silent=True) or {}

    # --- Input validation ---
    topic = _cf_sanitize_topic(data.get('topic', 'Research'))

    raw_sources = data.get('sources', [])
    if not isinstance(raw_sources, list):
        raw_sources = []
    # Enforce count limit before any processing (prevents large-array memory amplification)
    raw_sources = raw_sources[:_CF_MAX_SOURCES]
    sources = []
    for src in raw_sources:
        if not isinstance(src, str):
            continue
        if src.startswith('https://'):
            sources.append(src[:_CF_MAX_SOURCE_LENGTH])
        elif not src.startswith('http'):
            # Plain-text source — trim to safe length
            sources.append(src[:_CF_MAX_SOURCE_LENGTH])
        # http:// (non-TLS) URLs are silently dropped

    raw_types = data.get('artifact_types', list(_CF_ALLOWED_ARTIFACT_TYPES))
    if not isinstance(raw_types, list):
        raw_types = list(_CF_ALLOWED_ARTIFACT_TYPES)
    artifact_types = set(raw_types) & _CF_ALLOWED_ARTIFACT_TYPES
    if not artifact_types:
        artifact_types = set(_CF_ALLOWED_ARTIFACT_TYPES)

    auth_json = os.environ.get('NOTEBOOKLM_AUTH_JSON')
    if not auth_json:
        return flask.jsonify({'status': 'error', 'error': 'Configuration error'}), 500

    GCS_BUCKET = os.environ.get('GCS_RESEARCH_BUCKET', 'research-artifacts-n8n-2026')

    async def run():
        from notebooklm import NotebookLMClient
        from google.cloud import storage as gcs_storage

        # Write auth JSON to a temp file with restricted permissions (0o600).
        # NamedTemporaryFile default creates files readable only by owner on most systems,
        # but we explicitly set 0o600 to guarantee it regardless of umask.
        tmp_auth_path = None
        try:
            fd = os.open(
                tempfile.mktemp(suffix='.json'),  # noqa: S306 — path only, opened immediately
                os.O_WRONLY | os.O_CREAT | os.O_EXCL,
                0o600,
            )
            with os.fdopen(fd, 'w') as f:
                f.write(auth_json)
                tmp_auth_path = f.name
        except Exception:
            # Fallback: use NamedTemporaryFile and immediately restrict permissions
            with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
                f.write(auth_json)
                tmp_auth_path = f.name
            os.chmod(tmp_auth_path, 0o600)

        result = {
            'status': 'success',
            'sources_added': 0,
            'summary': None,
            'report_text': None,
            'mindmap_json': None,
            'mindmap_summary': None,
            'mindmap_url': None,
            'audio_url': None,
        }
        notebook_id = None

        try:
            async with await NotebookLMClient.from_storage(path=tmp_auth_path) as client:
                try:
                    nb = await client.notebooks.create(f'Research: {topic}')
                    notebook_id = nb.id

                    for src in sources:
                        try:
                            if src.startswith('https://'):
                                await client.sources.add_url(notebook_id, src, wait=True)
                            else:
                                await client.sources.add_text(notebook_id, 'Pasted content', src)
                            result['sources_added'] += 1
                        except Exception:
                            print(f'[research] Source add failed (index {result["sources_added"]})')

                    try:
                        result['summary'] = await client.notebooks.get_summary(notebook_id)
                    except Exception:
                        pass

                    gcs = gcs_storage.Client()
                    bucket = gcs.bucket(GCS_BUCKET)

                    # --- Report (fast, ~30s) ---
                    if 'report' in artifact_types:
                        tmp_report = None
                        try:
                            gen = await client.artifacts.generate_report(notebook_id)
                            await client.artifacts.wait_for_completion(
                                notebook_id, gen.task_id, timeout=120, initial_interval=5)
                            with tempfile.NamedTemporaryFile(suffix='.md', delete=False) as f:
                                tmp_report = f.name
                            await client.artifacts.download_report(notebook_id, tmp_report)
                            with open(tmp_report, 'r', encoding='utf-8') as f:
                                result['report_text'] = f.read()
                        except Exception:
                            print(f'[research] Report failed:\n{_tb.format_exc()}')
                            result['report_text'] = '[Report generation failed. Check function logs.]'
                        finally:
                            if tmp_report and os.path.exists(tmp_report):
                                os.unlink(tmp_report)

                    # --- Mind map — generate_mind_map returns dict directly ---
                    if 'mindmap' in artifact_types:
                        try:
                            mindmap = await client.artifacts.generate_mind_map(notebook_id)
                            result['mindmap_json'] = mindmap
                            result['mindmap_summary'] = _summarize_mindmap(mindmap)
                            blob = bucket.blob(f'research/{notebook_id}/mindmap.json')
                            blob.upload_from_string(
                                _json.dumps(mindmap, indent=2, ensure_ascii=False),
                                content_type='application/json',
                            )
                            result['mindmap_url'] = blob.generate_signed_url(
                                expiration=timedelta(hours=1), method='GET')
                        except Exception:
                            print(f'[research] Mindmap failed:\n{_tb.format_exc()}')
                            result['mindmap_summary'] = '[Mind map generation failed. Check function logs.]'

                    # --- Audio overview (slow, 2-5 min) — generated last ---
                    if 'audio' in artifact_types:
                        tmp_audio = None
                        try:
                            gen = await client.artifacts.generate_audio(
                                notebook_id,
                                instructions=f'Engaging deep-dive podcast about: {topic}',
                            )
                            await client.artifacts.wait_for_completion(
                                notebook_id, gen.task_id, timeout=600, initial_interval=10)
                            with tempfile.NamedTemporaryFile(suffix='.mp4', delete=False) as f:
                                tmp_audio = f.name
                            await client.artifacts.download_audio(notebook_id, tmp_audio)
                            blob = bucket.blob(f'research/{notebook_id}/audio.mp4')
                            blob.upload_from_filename(tmp_audio, content_type='audio/mp4')
                            result['audio_url'] = blob.generate_signed_url(
                                expiration=timedelta(hours=1), method='GET')
                        except Exception:
                            print(f'[research] Audio failed:\n{_tb.format_exc()}')
                            result['audio_error'] = 'Audio generation failed. Check function logs.'
                        finally:
                            if tmp_audio and os.path.exists(tmp_audio):
                                os.unlink(tmp_audio)

                finally:
                    if notebook_id:
                        try:
                            await client.notebooks.delete(notebook_id)
                        except Exception:
                            pass
        finally:
            # Always remove the auth credentials temp file
            if tmp_auth_path and os.path.exists(tmp_auth_path):
                os.unlink(tmp_auth_path)

        return result

    try:
        return flask.jsonify(asyncio.run(run())), 200
    except Exception:
        print(f'[research] Fatal error:\n{traceback.format_exc()}')
        return flask.jsonify({'status': 'error', 'error': 'Internal server error'}), 500


# ---------------------------------------------------------------------------
# CV → PDF Generator
# ---------------------------------------------------------------------------

@functions_framework.http
def generate_cv_pdf(request):
    """
    Convert a structured CV JSON into a clean PDF.

    POST body:
        cv_json  (dict)  — structured CV with keys: summary, skills, experience,
                           projects, education
        job_title (str)  — role name (used in PDF header)
        company   (str)  — company name (used in PDF header)

    Returns:
        JSON { pdf_base64: str, file_name: str }
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
    except ImportError:
        return flask.jsonify({'status': 'error', 'error': 'reportlab not installed'}), 500

    data = request.get_json(silent=True) or {}
    cv   = data.get('cv_json', {})
    if isinstance(cv, str):
        import json as _json
        try:
            cv = _json.loads(cv)
        except Exception:
            cv = {}

    job_title = data.get('job_title', 'Position')
    company   = data.get('company', '')

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm,  bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    H1  = ParagraphStyle('H1',  parent=styles['Heading1'], fontSize=20, textColor=colors.HexColor('#1a1a2e'), spaceAfter=4)
    H2  = ParagraphStyle('H2',  parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#4a90e2'), spaceBefore=12, spaceAfter=4, borderPad=2)
    SUB = ParagraphStyle('SUB', parent=styles['Normal'],   fontSize=10, textColor=colors.HexColor('#555555'), spaceAfter=8)
    NOR = ParagraphStyle('NOR', parent=styles['Normal'],   fontSize=10, leading=14, spaceAfter=4)
    BUL = ParagraphStyle('BUL', parent=styles['Normal'],   fontSize=9,  leading=13, leftIndent=12, spaceAfter=2)

    story = []

    # Header
    name = cv.get('summary', '')[:30] or 'Fabio Silva'  # fallback
    story.append(Paragraph('Fabio Silva', H1))
    story.append(Paragraph('AI Automation Engineer', SUB))
    if job_title:
        story.append(Paragraph(f'Application for: <b>{job_title}</b>' + (f' at {company}' if company else ''), SUB))
    story.append(Spacer(1, 0.3*cm))

    # Summary
    summary = cv.get('summary', '')
    if summary:
        story.append(Paragraph('Profile', H2))
        story.append(Paragraph(summary, NOR))

    # Skills
    skills = cv.get('skills', [])
    if skills:
        story.append(Paragraph('Skills', H2))
        # 2-column grid
        rows  = [skills[i:i+2] for i in range(0, len(skills), 2)]
        tdata = [[f'• {r[0]}', f'• {r[1]}' if len(r) > 1 else ''] for r in rows]
        t = Table(tdata, colWidths=[8.5*cm, 8.5*cm])
        t.setStyle(TableStyle([('FONTSIZE', (0,0), (-1,-1), 9), ('TOPPADDING', (0,0), (-1,-1), 2), ('BOTTOMPADDING', (0,0), (-1,-1), 2)]))
        story.append(t)

    # Experience
    experience = cv.get('experience', [])
    if experience:
        story.append(Paragraph('Experience', H2))
        for exp in experience:
            role    = exp.get('role', '')
            co      = exp.get('company', '')
            period  = exp.get('period', '')
            bullets = exp.get('bullets', [])
            story.append(Paragraph(f'<b>{role}</b> — {co} <font color="#888888">({period})</font>', NOR))
            for b in bullets:
                story.append(Paragraph(f'• {b}', BUL))
            story.append(Spacer(1, 0.2*cm))

    # Projects
    projects = cv.get('projects', [])
    if projects:
        story.append(Paragraph('Projects', H2))
        for p in projects:
            name_p = p.get('name', '')
            desc   = p.get('description', '')
            techs  = ', '.join(p.get('technologies', []))
            story.append(Paragraph(f'<b>{name_p}</b>' + (f' — {desc}' if desc else ''), NOR))
            if techs:
                story.append(Paragraph(f'<font color="#666666">Tech: {techs}</font>', BUL))

    # Education
    education = cv.get('education', [])
    if education:
        story.append(Paragraph('Education', H2))
        for e in education:
            story.append(Paragraph(f'<b>{e.get("degree","")}</b> — {e.get("institution","")} ({e.get("year","")})', NOR))

    doc.build(story)
    pdf_bytes  = buffer.getvalue()
    pdf_base64 = base64.b64encode(pdf_bytes).decode()
    file_name  = f"CV_{job_title.replace(' ','_')[:40]}.pdf"

    return flask.jsonify({'status': 'success', 'pdf_base64': pdf_base64, 'file_name': file_name}), 200
